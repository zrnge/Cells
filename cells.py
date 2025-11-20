import tkinter as tk
from tkinter import filedialog, ttk, messagebox, simpledialog
from openpyxl import load_workbook, Workbook
import csv
import os
from collections import deque
import re 
import subprocess # Needed to open links for documentation

# --- Tooltip Class (UNCHANGED) ---
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.id = None
        self.x = self.y = 0
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hide()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(500, self.show)

    def unschedule(self):
        id_ = self.id
        self.id = None
        if id_:
            self.widget.after_cancel(id_)

    def show(self):
        if self.tip_window or not self.text: return
        x, y, cx, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 25
        y = y + cy + self.widget.winfo_rooty() + 25
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        self.tip_window.attributes('-alpha', 0.9)
        label = tk.Label(self.tip_window, text=self.text, justify=tk.LEFT,
                         background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hide(self):
        if self.tip_window:
            self.tip_window.destroy()
        self.tip_window = None

# --- Custom Paste Position Dialog (UNCHANGED) ---
class PastePositionDialog(simpledialog.Dialog):
    def __init__(self, parent, title, orientation, current_index, max_index):
        self.orientation = orientation 
        self.current_index = current_index
        self.max_index = max_index
        self.position = None
        super().__init__(parent, title=title)

    def body(self, master):
        
        if self.orientation == 'row':
            label_text = "Select where to paste the new data:"
            append_text = "Append as New Rows (to end)"
            index_label = f"Target Row (1-based): {self.current_index + 1}" if self.current_index is not None else "Target: None Selected"
        else: 
            label_text = "Select where to paste the new data:"
            append_text = "Append as New Columns (to end)"
            index_label = f"Target Column (1-based): {self.current_index + 1}" if self.current_index is not None else "Target: None Selected"

        tk.Label(master, text=label_text).pack(pady=5)
        tk.Label(master, text=index_label, fg='yellow').pack(pady=2)

        self.pos_var = tk.StringVar(master, value="OVERWRITE_START")
        
        opt_frame = tk.Frame(master)
        opt_frame.pack(pady=10)

        tk.Radiobutton(opt_frame, text=append_text, variable=self.pos_var, value="APPEND").pack(anchor=tk.W)

        if self.current_index is not None:
            tk.Radiobutton(opt_frame, text=f"Overwrite starting at current {self.orientation}", variable=self.pos_var, value="OVERWRITE_START").pack(anchor=tk.W)
            tk.Radiobutton(opt_frame, text=f"Insert New {self.orientation.title()} Before", variable=self.pos_var, value="INSERT_BEFORE").pack(anchor=tk.W)
            tk.Radiobutton(opt_frame, text=f"Insert New {self.orientation.title()} After", variable=self.pos_var, value="INSERT_AFTER").pack(anchor=tk.W)

        return None

    def apply(self):
        self.position = self.pos_var.get()

# --- Custom Sheet Rename Dialog (UNCHANGED) ---
class RenameSheetDialog(simpledialog.Dialog):
    def __init__(self, parent, title, sheet_names):
        self.sheet_names = sheet_names
        self.selected_sheet = None
        self.new_name = None
        super().__init__(parent, title=title)

    def body(self, master):
        tk.Label(master, text="Select Sheet to Rename:").pack(pady=5)
        
        self.sheet_var = tk.StringVar(master)
        self.sheet_combo = ttk.Combobox(master, textvariable=self.sheet_var, values=self.sheet_names, state='readonly')
        self.sheet_combo.pack(pady=5)
        if self.sheet_names:
            self.sheet_combo.set(self.sheet_names[0])

        tk.Label(master, text="New Sheet Name:").pack(pady=5)
        self.name_entry = tk.Entry(master)
        self.name_entry.pack(pady=5)
        return self.sheet_combo 

    def apply(self):
        self.selected_sheet = self.sheet_var.get()
        self.new_name = self.name_entry.get().strip()
        if not self.new_name:
            messagebox.showerror("Error", "New sheet name cannot be empty.")
            self.result = None
            return

# ---------------------------------------------

class ExcelEditor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel/CSV Editor")
        self.root.geometry("1000x600")

        if os.path.exists("icon.ico"):
            self.root.iconbitmap("icon.ico")

        self.bg_color = "#2e2e2e"
        self.fg_color = "#ffffff"
        self.highlight_color = "#007acc"
        self.button_bg = "#3e3e3e"
        self.button_fg = "#ffffff"

        self.root.configure(bg=self.bg_color)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close) 

        self.data_rows = []
        self.file_path = None
        self.file_type = None
        self.workbook = None 
        self.sheet_names = [] 
        self.current_sheet = None 
        self.unsaved_changes = False
        self.history = deque(maxlen=50) 
        self.history_index = -1
        self.is_undoing = False
        
        self.selected_row_index = None 
        self.selected_col_index = None 
        self.selected_row = None
        self.selected_cell_value = None
        
        self.current_sort_col = None 
        self.current_sort_reverse = False
        
        self.show_grid = tk.BooleanVar(value=True) 

        self._configure_styles()
        self._create_menu_bar()
        self._create_icon_bar()
        self._create_widgets()
        
        self.history.clear()
        self._save_state()
        self.unsaved_changes = False
        self._update_status_bar()

    def _configure_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure("Treeview", 
                        background="#1e1e1e", 
                        foreground=self.fg_color,
                        fieldbackground="#1e1e1e",
                        rowheight=25,
                        bordercolor=self.bg_color,
                        borderwidth=0) 
                        
        style.map("Treeview", background=[("selected", self.highlight_color)],
                  foreground=[("selected", self.fg_color)])
                  
        style.configure("Treeview.Heading", background="#3e3e3e", foreground="#ffffff", relief="flat")
        style.map("Treeview.Heading", background=[('active', '#5e5e5e')])
        
        style.configure("TButton", background=self.button_bg, foreground=self.button_fg, padding=(2, 2))
        style.configure('TCombobox', fieldbackground=self.button_bg, foreground=self.button_fg, 
                        selectbackground=self.highlight_color, selectforeground=self.fg_color)
                        
        style.configure("Grid.Treeview", 
                        bordercolor="#555555", 
                        borderwidth=1)

    # ---------------------------------------------
    # --- Help Menu Logic ---
    def show_documentation(self):
        """Attempts to open the local README.md or the GitHub repository."""
        doc_path = os.path.join(os.path.dirname(__file__), "README.md")
        if os.path.exists(doc_path):
            try:
                # Use subprocess to open the file with the default viewer/editor
                os.startfile(doc_path)
            except AttributeError:
                # Handle non-Windows systems
                subprocess.call(('open', doc_path))
            except Exception:
                messagebox.showinfo("Documentation", f"Could not open local README.md. Please open it manually from: {doc_path}")
        else:
            messagebox.showinfo("Documentation", "Local documentation (README.md) not found. Checking GitHub...")
            self._open_external_link("https://github.com/zrnge")
            
    def show_about(self):
        """Displays information about the zrnge project."""
        about_text = (
            "Excel/CSV Editor\n\n"
            "Developed by zrnge\n"
            "Project: https://github.com/zrnge\n\n"
            "zrnge is a developer focused on creating fast, efficient, "
            "and user-friendly tools. This editor is a demonstration "
            "of building powerful GUI applications with Python's standard "
            "libraries and minimal dependencies."
        )
        messagebox.showinfo("About Excel/CSV Editor", about_text)

    def _open_external_link(self, url):
        """Helper to open a link in the default browser."""
        import webbrowser
        webbrowser.open_new_tab(url)
    
    # ---------------------------------------------

    def toggle_grid_lines(self):
        """Toggles the visibility of grid lines by changing the Treeview style."""
        if self.show_grid.get():
            self.tree.configure(style="Grid.Treeview")
        else:
            self.tree.configure(style="Treeview")
            
    def _create_menu_bar(self):
        menubar = tk.Menu(self.root, bg=self.bg_color, fg=self.fg_color)
        self.root.config(menu=menubar)

        # --- File Menu ---
        file_menu = tk.Menu(menubar, tearoff=0, bg=self.bg_color, fg=self.fg_color)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="New Sheet", command=self.create_new_sheet)
        file_menu.add_command(label="Open...", command=self.open_file)
        file_menu.add_separator()
        file_menu.add_command(label="Save", command=self.save_file)
        file_menu.add_command(label="Save As...", command=self.save_as_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self._on_close)

        # --- Edit Menu ---
        edit_menu = tk.Menu(menubar, tearoff=0, bg=self.bg_color, fg=self.fg_color)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Undo", command=self.undo)
        edit_menu.add_command(label="Redo", command=self.redo)
        edit_menu.add_separator()
        edit_menu.add_command(label="Add Row", command=self.add_row)
        edit_menu.add_command(label="Add Column", command=self.add_column)
        edit_menu.add_command(label="Rename Sheet...", command=self.rename_sheet) 
        edit_menu.add_separator()
        edit_menu.add_command(label="📋 Paste Vertical (into Column)", command=self.paste_vertical) 
        edit_menu.add_command(label="➡️ Paste Horizontal (into Row)", command=self.paste_horizontal) 

        # --- View Menu ---
        view_menu = tk.Menu(menubar, tearoff=0, bg=self.bg_color, fg=self.fg_color)
        menubar.add_cascade(label="View", menu=view_menu)
        view_menu.add_checkbutton(label="Show Grid Lines", 
                                  variable=self.show_grid, 
                                  command=self.toggle_grid_lines)

        # --- NEW: Help Menu ---
        help_menu = tk.Menu(menubar, tearoff=0, bg=self.bg_color, fg=self.fg_color)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="Documentation", command=self.show_documentation)
        help_menu.add_command(label="About", command=self.show_about)


    def _create_icon_bar(self):
        icon_bar = tk.Frame(self.root, bg="#1e1e1e") 
        icon_bar.pack(fill=tk.X, padx=0, pady=0)

        icons_left = [
            ("📁", "Open File", self.open_file),
            ("💾", "Save File", self.save_file),
            ("↩️", "Undo", self.undo),
            ("↪️", "Redo", self.redo),
            ("➕R", "Add Row", self.add_row),
            ("➕C", "Add Column", self.add_column),
        ]
        
        for text, tooltip, cmd in icons_left:
            btn = tk.Button(icon_bar, text=text, command=cmd, bg=self.button_bg, fg=self.button_fg, 
                            width=5, relief=tk.FLAT, activebackground=self.highlight_color, activeforeground=self.button_fg) 
            btn.pack(side=tk.LEFT, padx=1, pady=1)
            Tooltip(btn, tooltip) 

        sheet_label = tk.Label(icon_bar, text="Sheet:", bg="#1e1e1e", fg=self.fg_color)
        sheet_label.pack(side=tk.LEFT, padx=(10, 2), pady=1)
        
        self.sheet_selector = ttk.Combobox(icon_bar, state="readonly", width=20, values=self.sheet_names)
        self.sheet_selector.pack(side=tk.LEFT, padx=2, pady=1)
        self.sheet_selector.bind("<<ComboboxSelected>>", self.switch_sheet)
        self.sheet_selector.set("No Sheets Loaded")

        search_frame = tk.Frame(icon_bar, bg="#1e1e1e")
        search_frame.pack(side=tk.RIGHT, padx=5, pady=1)

        self.search_entry = tk.Entry(search_frame, bg=self.button_bg, fg=self.fg_color, insertbackground=self.fg_color, width=40)
        self.search_entry.pack(side=tk.LEFT, padx=2, pady=2, fill=tk.X, expand=True)
        self.search_entry.insert(0, "Search (e.g. key or Col:val1,val2)")
        self.search_entry.bind("<FocusIn>", self._clear_placeholder)
        self.search_entry.bind("<FocusOut>", self._restore_placeholder)
        self.search_entry.bind("<Return>", self.apply_search_filter) 

        search_btn = tk.Button(search_frame, text="🔎", command=self.apply_search_filter,
                  bg=self.button_bg, fg=self.button_fg, width=3, relief=tk.FLAT)
        search_btn.pack(side=tk.LEFT, padx=1, pady=1)
        Tooltip(search_btn, "Apply Search Filter")
        
        clear_btn = tk.Button(search_frame, text="🧹", command=self.clear_filter,
                  bg=self.button_bg, fg=self.button_fg, width=3, relief=tk.FLAT)
        clear_btn.pack(side=tk.LEFT, padx=1, pady=1)
        Tooltip(clear_btn, "Clear Filter")

    def _clear_placeholder(self, event):
        if self.search_entry.get() == "Search (e.g. key or Col:val1,val2)":
            self.search_entry.delete(0, tk.END)
            self.search_entry.config(fg=self.fg_color)
            
    def _restore_placeholder(self, event):
        if not self.search_entry.get():
            self.search_entry.insert(0, "Search (e.g. key or Col:val1,val2)")
            self.search_entry.config(fg='gray') 

    def _create_widgets(self):
        self.frame = tk.Frame(self.root, bg=self.bg_color)
        self.frame.pack(fill=tk.BOTH, expand=True)

        self.vsb = tk.Scrollbar(self.frame, orient="vertical")
        self.vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.hsb = tk.Scrollbar(self.frame, orient="horizontal")
        self.hsb.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree = ttk.Treeview(self.frame, yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.vsb.config(command=self.tree.yview)
        self.hsb.config(command=self.tree.xview)
        
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<Double-1>", self.edit_cell, add="+")
        self.tree.bind("<Button-1>", self.handle_header_click) 
        self.tree.bind("<Double-1>", self.handle_header_double_click, add="+")
        
        self._create_context_menu()
        
        self.status_bar = tk.Label(self.root, text="Ready: No file loaded.", 
                                   bd=1, relief=tk.SUNKEN, anchor=tk.W, 
                                   bg=self.bg_color, fg=self.fg_color)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.toggle_grid_lines() 

    def _create_context_menu(self):
        self.menu = tk.Menu(self.root, tearoff=0, bg=self.bg_color, fg=self.fg_color)
        self.menu.add_command(label="Copy Cell", command=self.copy_cell)
        self.menu.add_command(label="Copy Row", command=self.copy_row)
        self.menu.add_command(label="Copy Column", command=self.copy_column)
        self.menu.add_separator()
        
        self.menu.add_command(label="📋 Paste Vertical (into Column)", command=self.paste_vertical) 
        self.menu.add_command(label="➡️ Paste Horizontal (into Row)", command=self.paste_horizontal) 
        self.menu.add_separator()

        self.menu.add_command(label="Add Row Above", command=self.add_row_above)
        self.menu.add_command(label="Add Row Below", command=self.add_row_below)
        self.menu.add_command(label="Delete Row", command=self.delete_row)
        self.menu.add_command(label="Move Row Up", command=self.move_row_up)
        self.menu.add_command(label="Move Row Down", command=self.move_row_down)
        self.menu.add_command(label="Delete Column", command=self.delete_column)
        self.menu.add_command(label="Move Column Left", command=self.move_column_left)
        self.menu.add_command(label="Move Column Right", command=self.move_column_right)
        self.menu.add_command(label="Clear Cell", command=self.clear_cell)
        self.menu.add_separator()
        self.menu.add_command(label="Search (from search box)", command=self.apply_search_filter)

    # ---------------- State Management / Undo/Redo (UNCHANGED) ----------------
    def _save_state(self):
        if self.is_undoing: return
        if self.history_index < len(self.history) - 1:
            while len(self.history) > self.history_index + 1:
                self.history.pop()
        current_state = ([list(row) for row in self.data_rows], list(self.tree["columns"]))
        self.history.append(current_state)
        self.history_index = len(self.history) - 1
        self.unsaved_changes = True
        self._update_status_bar()

    def _load_state(self, index):
        if index < 0 or index >= len(self.history): return
        self.is_undoing = True
        self.history_index = index
        self.data_rows, columns = self.history[index]
        self.tree["columns"] = columns
        self._refresh_headings()
        self.clear_filter() 
        self.is_undoing = False
        self.unsaved_changes = (index != len(self.history) - 1)
        self._update_status_bar()

    def undo(self):
        if self.history_index > 0:
            self._load_state(self.history_index - 1)
        
    def redo(self):
        if self.history_index < len(self.history) - 1:
            self._load_state(self.history_index + 1)
            
    # ---------------- Status Bar / Exit (UNCHANGED) ----------------
    def _update_status_bar(self, message=None):
        if message:
            self.status_bar.config(text=message)
            return
        file_path = self.file_path if self.file_path is not None else '' 
        file_name = os.path.basename(file_path)
        file_name = file_name if file_name else 'None'
        sheet_info = f" | Sheet: {self.current_sheet}" if self.current_sheet else ""
        row_count = len(self.data_rows)
        col_count = len(self.tree["columns"]) if self.tree["columns"] else 0
        status_text = f"File: {file_name}{sheet_info} | Rows: {row_count} | Columns: {col_count}"
        window_title = "Excel/CSV Editor"
        if file_name and file_name != 'None':
            window_title += f" - {file_name}"
        if self.current_sheet:
            window_title += f" ({self.current_sheet})"
        if self.unsaved_changes:
            status_text += " | **UNSAVED CHANGES**"
            window_title += " *"
        self.status_bar.config(text=status_text)
        self.root.title(window_title)

    def _on_close(self):
        if self.unsaved_changes:
            response = messagebox.askyesnocancel("Unsaved Changes", 
                                                "You have unsaved changes. Do you want to save before exiting?")
            if response is None: return
            elif response is True: 
                self.save_file()
                if self.unsaved_changes: return
        self.root.destroy()
        
    # ---------------- File/Sheet Loading/Saving ----------------
    def open_file(self):
        if self.unsaved_changes:
            if not messagebox.askyesno("Unsaved Changes", "Discard unsaved changes and open a new file?"):
                return
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv")])
        if not file_path: return
        self.file_path = file_path
        self.unsaved_changes = False
        self.workbook = None 
        try:
            if file_path.lower().endswith((".xlsx", ".xls")):
                self.file_type = "excel"
                self.workbook = load_workbook(filename=file_path, data_only=True)
                self.sheet_names = self.workbook.sheetnames
                self.current_sheet = self.sheet_names[0]
                self.sheet_selector.config(values=self.sheet_names)
                self.sheet_selector.set(self.current_sheet)
                self.read_excel_sheet(self.current_sheet)
            elif file_path.lower().endswith(".csv"):
                self.file_type = "csv"
                self.workbook = None # CSV doesn't use openpyxl Workbook
                self.sheet_names = ["Data"] 
                self.current_sheet = "Data"
                self.sheet_selector.config(values=self.sheet_names)
                self.sheet_selector.set(self.current_sheet)
                self.read_csv(file_path)
            self.history.clear()
            self.current_sort_col = None
            self.current_sort_reverse = False
            self._save_state()
            self.unsaved_changes = False 
            self._update_status_bar()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open file\n{e}")
            self.file_path = None 

    def switch_sheet(self, event):
        if self.file_type != "excel" or self.workbook is None: return
        new_sheet_name = self.sheet_selector.get()
        if new_sheet_name == self.current_sheet: return
        if self.unsaved_changes:
            if not messagebox.askyesno("Unsaved Changes", "Switching sheets will discard unsaved changes. Continue?"):
                self.sheet_selector.set(self.current_sheet) 
                return
        self.current_sheet = new_sheet_name
        self.read_excel_sheet(self.current_sheet)
        self.history.clear()
        self.current_sort_col = None
        self.current_sort_reverse = False
        self._save_state()
        self.unsaved_changes = False 
        self._update_status_bar()

    def read_excel_sheet(self, sheet_name):
        sheet = self.workbook[sheet_name]
        self.tree.delete(*self.tree.get_children())
        headers = [str(cell.value) if cell.value is not None else "" for cell in next(sheet.iter_rows(max_row=1))]
        self.tree["columns"] = headers
        self.tree["show"] = "headings"
        self._refresh_headings()
        self.data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data_row = [str(v) if v is not None else "" for v in row]
            data_row = data_row[:len(headers)] + [""] * (len(headers) - len(data_row))
            self.data_rows.append(data_row)
            self.tree.insert("", "end", values=data_row)

    def read_csv(self, file_path):
        with open(file_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows: return
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = rows[0]
        self.tree["show"] = "headings"
        self._refresh_headings()
        self.data_rows = [list(row) for row in rows[1:]]
        for row in self.data_rows:
            self.tree.insert("", "end", values=row)

    def save_file(self):
        if not self.file_path:
            self.save_as_file()
            return
        if self.file_type == "excel" and not self.workbook:
             messagebox.showerror("Error", "Workbook object not loaded. Please use 'Save As'.")
             return
        self._save_to_file(self.file_path)

    def save_as_file(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx" if self.file_type == "excel" else ".csv",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")],
        )
        if file_path:
            self.file_path = file_path
            self._save_to_file(file_path)

    def _save_to_file(self, file_path):
        try:
            data_to_save = self.data_rows 
            headers = self.tree["columns"]
            
            if file_path.lower().endswith((".xlsx", ".xls")) and self.file_type == "excel":
                if not self.workbook: 
                    self.workbook = Workbook()
                    ws = self.workbook.active
                    ws.title = self.current_sheet if self.current_sheet else "Sheet1"
                else:
                    ws = self.workbook[self.current_sheet]

                ws.delete_rows(1, ws.max_row) 
                ws.append(headers)
                for row in data_to_save:
                    ws.append(row)
                    
                self.workbook.save(file_path)
                
            elif file_path.lower().endswith(".csv") or self.file_type == "csv":
                with open(file_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(data_to_save)
                    
            self.unsaved_changes = False
            self._update_status_bar()
            messagebox.showinfo("Saved", f"File saved successfully: {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file\n{e}")

    def rename_sheet(self):
        if self.file_type != "excel" or not self.workbook:
            messagebox.showwarning("Warning", "Sheet renaming is only available for open Excel files or new sheets.")
            return

        dialog = RenameSheetDialog(self.root, "Rename Sheet", self.sheet_names)
        
        selected_sheet = dialog.selected_sheet
        new_name = dialog.new_name

        if not selected_sheet or not new_name:
            return

        if new_name in self.sheet_names:
            messagebox.showerror("Error", f"Sheet name '{new_name}' already exists.")
            return

        try:
            sheet = self.workbook[selected_sheet]
            sheet.title = new_name
            
            old_index = self.sheet_names.index(selected_sheet)
            self.sheet_names[old_index] = new_name
            
            self.sheet_selector.config(values=self.sheet_names)
            
            if self.current_sheet == selected_sheet:
                self.current_sheet = new_name
                self.sheet_selector.set(new_name)
                self._update_status_bar()
            else:
                 self.sheet_selector.set(self.current_sheet)
            
            self.unsaved_changes = True
            self._update_status_bar(f"Sheet '{selected_sheet}' renamed to '{new_name}'.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to rename sheet: {e}")
            
    # ---------------- Data / Structure Manipulation ----------------
    def create_new_sheet(self):
        # 1. Clear Data and UI
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ["Column1", "Column2", "Column3"]
        self.tree["show"] = "headings"
        self._refresh_headings()
        self.data_rows = []
        
        # 2. Initialize a new Workbook (THE FIX)
        self.file_path = None
        self.file_type = "excel"
        self.workbook = Workbook()
        
        # Set initial sheet name and update UI
        ws = self.workbook.active
        ws.title = "Sheet1"
        self.sheet_names = ["Sheet1"]
        self.current_sheet = "Sheet1"
        self.sheet_selector.config(values=self.sheet_names)
        self.sheet_selector.set("Sheet1")
        
        # 3. Reset State
        self.history.clear()
        self.current_sort_col = None
        self.current_sort_reverse = False
        self._save_state()
        self.unsaved_changes = False
        self._update_status_bar()

    def _insert_new_column(self, col_index):
        col_name = simpledialog.askstring("New Column Name", f"Enter name for column #{col_index + 1}:")
        if not col_name: return False

        headers = list(self.tree["columns"])
        headers.insert(col_index, col_name)
        self.tree["columns"] = headers
        self._refresh_headings()
        
        for row in self.data_rows:
            if col_index > len(row):
                 row.extend([""] * (col_index - len(row)))
            row.insert(col_index, "")
        return True

    def _insert_new_row(self, row_index):
        new_row = [""] * len(self.tree["columns"])
        self.data_rows.insert(row_index, new_row)

    def paste_vertical(self):
        data_2d, delimiter = self._get_paste_data()
        if data_2d is None or not data_2d: return

        data_list = [item for sublist in data_2d for item in sublist]
        rows_to_paste = len(data_list)
        
        orientation = 'row'
        dialog = PastePositionDialog(self.root, "Paste Vertical Position", orientation, 
                                     self.selected_row_index, len(self.data_rows))
        
        position = dialog.position
        if position is None: return

        start_row = self.selected_row_index if self.selected_row_index is not None else len(self.data_rows)
        start_col = self.selected_col_index if self.selected_col_index is not None else 0
        
        if position == "INSERT_BEFORE":
            for _ in range(rows_to_paste): self._insert_new_row(start_row)
        elif position == "INSERT_AFTER":
            for _ in range(rows_to_paste): self._insert_new_row(start_row + 1)
            start_row += 1 
        elif position == "APPEND":
            start_row = len(self.data_rows)
            for _ in range(rows_to_paste): self._insert_new_row(len(self.data_rows))
        
        virtual_result = f"Pasting {rows_to_paste} cells vertically into Column {self.tree['columns'][start_col]}."
        virtual_result += f"\nStarting at Row {start_row + 1}. Mode: {position}."
        if not messagebox.askokcancel("Confirm Vertical Paste", virtual_result): return

        for i, value in enumerate(data_list):
            target_row_index = start_row + i
            target_row = self.data_rows[target_row_index]
            
            if start_col >= len(target_row):
                target_row.extend([""] * (start_col - len(target_row) + 1))
            
            target_row[start_col] = value

        self.clear_filter() 
        self._save_state()
        self._update_status_bar(f"Pasted {rows_to_paste} cells vertically.")


    def paste_horizontal(self):
        data_2d, delimiter = self._get_paste_data()
        if data_2d is None or not data_2d: return
        
        data_list = [item for sublist in data_2d for item in sublist]
        cols_to_paste = len(data_list)

        orientation = 'col'
        dialog = PastePositionDialog(self.root, "Paste Horizontal Position", orientation, 
                                     self.selected_col_index, len(self.tree["columns"]))
        
        position = dialog.position
        if position is None: return
        
        start_row = self.selected_row_index if self.selected_row_index is not None else 0
        start_col = self.selected_col_index if self.selected_col_index is not None else len(self.tree["columns"])
        
        if position == "INSERT_BEFORE":
            for _ in range(cols_to_paste): 
                if not self._insert_new_column(start_col): return
        elif position == "INSERT_AFTER":
            for _ in range(cols_to_paste): 
                if not self._insert_new_column(start_col + 1): return
            start_col += cols_to_paste
        elif position == "APPEND":
            start_col = len(self.tree["columns"])
            for _ in range(cols_to_paste): 
                if not self._insert_new_column(len(self.tree["columns"])): return

        current_col_name = self.tree['columns'][start_col] if start_col < len(self.tree['columns']) else "NEW"
        virtual_result = f"Pasting {cols_to_paste} cells horizontally into Row {start_row + 1}."
        virtual_result += f"\nStarting at Column {current_col_name}. Mode: {position}."

        if not messagebox.askokcancel("Confirm Horizontal Paste", virtual_result): return

        target_row = self.data_rows[start_row]
        
        if position in ("OVERWRITE_START", "APPEND"):
            if start_col + cols_to_paste > len(target_row):
                padding = start_col + cols_to_paste - len(target_row)
                target_row.extend([""] * padding)
                
        for i, value in enumerate(data_list):
            target_row[start_col + i] = value

        self.clear_filter() 
        self._save_state()
        self._update_status_bar(f"Pasted {cols_to_paste} cells horizontally.")

    # ---------------- Other Manipulation Functions (UNCHANGED logic) ----------------

    def add_row(self):
        if not self.tree["columns"]:
            messagebox.showwarning("Warning", "Open a file or create new sheet first.")
            return
        new_row = [""] * len(self.tree["columns"])
        self.data_rows.append(new_row)
        self.tree.insert("", "end", values=new_row)
        self._save_state()

    def add_column(self):
        if not self.tree["columns"]:
            messagebox.showwarning("Warning", "Open a file or create new sheet first.")
            return
        self._insert_new_column(len(self.tree["columns"]))
        self.clear_filter()
        self._save_state()

    def delete_row(self):
        if self.selected_row_index is not None:
            del self.data_rows[self.selected_row_index]
            self.tree.delete(self.tree.selection())
            self._save_state()

    def delete_column(self):
        if self.selected_col_index is not None:
            col_idx = self.selected_col_index
            headers = list(self.tree["columns"])
            headers.pop(col_idx)
            self.tree["columns"] = headers
            for row in self.data_rows:
                if col_idx < len(row):
                    row.pop(col_idx)
            self._refresh_headings()
            self.clear_filter()
            self._save_state()
            
    def add_row_above(self):
        if self.selected_row_index is None: return
        self._insert_new_row(self.selected_row_index)
        self.clear_filter()
        self._save_state()

    def add_row_below(self):
        if self.selected_row_index is None: return
        self._insert_new_row(self.selected_row_index + 1)
        self.clear_filter()
        self._save_state()

    def clear_cell(self):
        selected_item_id = self.tree.selection()
        if not selected_item_id: return
        col_index = self.selected_col_index
        item_values = list(self.tree.item(selected_item_id[0], "values"))
        
        if col_index is not None and col_index < len(item_values):
            item_values[col_index] = ""
            self.tree.item(selected_item_id[0], values=item_values) 
            try:
                for i, row in enumerate(self.data_rows):
                    if self.selected_row_index is not None and i == self.selected_row_index:
                        self.data_rows[i][col_index] = ""
                        self._save_state()
                        return
            except Exception:
                pass 
                
    def move_row_up(self):
        idx = self.selected_row_index
        if idx is None or idx == 0: return
        self.data_rows[idx - 1], self.data_rows[idx] = self.data_rows[idx], self.data_rows[idx - 1]
        self.clear_filter()
        self.tree.selection_set(self.tree.get_children()[idx-1])
        self._save_state()

    def move_row_down(self):
        idx = self.selected_row_index
        if idx is None or idx >= len(self.data_rows)-1: return
        self.data_rows[idx + 1], self.data_rows[idx] = self.data_rows[idx], self.data_rows[idx + 1]
        self.clear_filter()
        self.tree.selection_set(self.tree.get_children()[idx+1])
        self._save_state()

    def move_column_left(self):
        idx = self.selected_col_index
        if idx is None or idx == 0: return
        headers = list(self.tree["columns"])
        headers[idx-1], headers[idx] = headers[idx], headers[idx-1]
        self.tree["columns"] = headers
        for row in self.data_rows:
            if idx < len(row):
                 row[idx-1], row[idx] = row[idx], row[idx-1]
        self._refresh_headings()
        self.clear_filter()
        self._save_state()

    def move_column_right(self):
        idx = self.selected_col_index
        headers = list(self.tree["columns"])
        if idx is None or idx >= len(headers)-1: return
        headers[idx+1], headers[idx] = headers[idx], headers[idx+1]
        self.tree["columns"] = headers
        for row in self.data_rows:
            if idx + 1 < len(row):
                row[idx+1], row[idx] = row[idx], row[idx+1]
        self._refresh_headings()
        self.clear_filter()
        self._save_state()

    def _refresh_headings(self):
        for idx, col_name in enumerate(self.tree["columns"]):
            name_only = col_name.replace(' ▲', '').replace(' ▼', '')
            indicator = ''
            if idx == self.current_sort_col:
                indicator = ' ▼' if self.current_sort_reverse else ' ▲'
            self.tree.heading(f"#{idx+1}", text=name_only + indicator)
            self.tree.column(f"#{idx+1}", width=120, anchor="center")

    def handle_header_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "heading": return
        col = self.tree.identify_column(event.x)
        if not col or col == '#0': return
        col_index = int(col.replace("#", "")) - 1
        self.sort_by_column(col_index)

    def handle_header_double_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "heading": return
        col = self.tree.identify_column(event.x)
        col_index = int(col.replace("#", "")) - 1
        
        old_name = self.tree["columns"][col_index].replace(' ▲', '').replace(' ▼', '')
        new_name = simpledialog.askstring("Edit Column", "Enter new column name:", initialvalue=old_name)
        if new_name and new_name != old_name:
            columns = list(self.tree["columns"])
            columns[col_index] = new_name
            self.tree["columns"] = columns
            self._refresh_headings()
            self._save_state()
            
    def sort_by_column(self, col_index):
        if self.current_sort_col == col_index:
            self.current_sort_reverse = not self.current_sort_reverse
        else:
            self.current_sort_col = col_index
            self.current_sort_reverse = False
        def convert_value(val):
            try:
                if isinstance(val, (int, float)): return float(val)
                return float(str(val).strip())
            except (ValueError, TypeError):
                return str(val).lower()
        def sort_key(row):
            try:
                if col_index < len(row):
                    val = row[col_index]
                    return convert_value(val)
                return ""
            except IndexError:
                return ""
        self.data_rows.sort(key=sort_key, reverse=self.current_sort_reverse)
        self.clear_filter()
        self._refresh_headings()
        self._save_state() 
        
    def edit_cell(self, event):
        row_id = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not row_id or not col or col == '#0': return

        col_index = int(col.replace("#", "")) - 1
        x, y, width, height = self.tree.bbox(row_id, column=col)
        
        current_row_values = list(self.tree.item(row_id, "values"))
        try:
            master_row_index = [idx for idx, row in enumerate(self.data_rows) if all(str(row[j]) == str(current_row_values[j]) for j in range(len(current_row_values)))][0]
        except Exception:
            return 
        
        current_value = self.data_rows[master_row_index][col_index]

        self.edit_entry = tk.Entry(self.tree)
        self.edit_entry.place(x=x, y=y, width=width, height=height)
        self.edit_entry.insert(0, current_value)
        self.edit_entry.focus()

        def update_visuals(event=None):
            new_value = self.edit_entry.get()
            temp_row = list(current_row_values)
            temp_row[col_index] = new_value
            self.tree.item(row_id, values=temp_row)

        def finalize_edit(event=None):
            if not self.edit_entry.winfo_exists(): return
            new_value = self.edit_entry.get()
            self.edit_entry.destroy() 
            
            if new_value != current_value:
                self.data_rows[master_row_index][col_index] = new_value
                self.tree.item(row_id, values=self.data_rows[master_row_index]) 
                self._save_state()

        self.edit_entry.bind("<KeyRelease>", update_visuals)
        self.edit_entry.bind("<Return>", finalize_edit)
        self.edit_entry.bind("<FocusOut>", finalize_edit)

    def show_context_menu(self, event):
        row_id = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        
        self.selected_row_index = None 
        self.selected_col_index = None
        self.selected_row = None
        self.selected_cell_value = None
        self.tree.selection_remove(self.tree.selection())

        if row_id and col and col != '#0':
            self.selected_row = self.tree.item(row_id)["values"]
            self.selected_col_index = int(col.replace("#", "")) - 1
            self.selected_cell_value = self.selected_row[self.selected_col_index]
            
            try:
                search_row_str = [str(v) for v in self.selected_row]
                self.selected_row_index = [idx for idx, row in enumerate(self.data_rows) if [str(v) for v in row] == search_row_str][0]
            except Exception:
                pass 

            self.tree.selection_set(row_id)
            self.menu.post(event.x_root, event.y_root)

    def copy_cell(self):
        if self.selected_cell_value is not None:
            self.root.clipboard_clear()
            self.root.clipboard_append(str(self.selected_cell_value))
            self._update_status_bar(f"Cell copied: {self.selected_cell_value}")

    def copy_row(self):
        if self.selected_row:
            self.root.clipboard_clear()
            self.root.clipboard_append("\t".join(map(str, self.selected_row)))
            self._update_status_bar("Row copied to clipboard!")

    def copy_column(self):
        if self.selected_col_index is not None:
            col_data = [str(row[self.selected_col_index]) for row in self.data_rows if self.selected_col_index < len(row)]
            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(col_data))
            self._update_status_bar("Column copied to clipboard!")
            
    def apply_search_filter(self, event=None):
        query = self.search_entry.get().strip()
        if not query or query == "Search (e.g. key or Col:val1,val2)":
            self.clear_filter()
            return
        
        filtered_rows = []
        
        if ":" in query:
            parts = query.split(":", 1)
            if len(parts) < 2: 
                messagebox.showwarning("Search Format", "Invalid column search format. Use Column:value1,value2,...")
                self.clear_filter()
                return

            col_name, values = parts
            col_name = col_name.strip()
            values_list = [v.strip().lower() for v in values.split(",") if v.strip()]
            
            try:
                col_index = list(self.tree["columns"]).index(col_name)
            except ValueError:
                messagebox.showwarning("Warning", f"Column '{col_name}' not found.")
                self.clear_filter()
                return
            
            for row in self.data_rows:
                cell = row[col_index] if col_index < len(row) else ""
                if cell and any(val in str(cell).lower() for val in values_list):
                    filtered_rows.append(row)
        else:
            value = query.lower()
            for row in self.data_rows:
                if any(cell and value in str(cell).lower() for cell in row):
                    filtered_rows.append(row)

        self.tree.delete(*self.tree.get_children())
        for row in filtered_rows:
            self.tree.insert("", "end", values=row)
        
        self._update_status_bar(f"Filter applied. {len(filtered_rows)} of {len(self.data_rows)} rows shown.")

    def clear_filter(self):
        self.tree.delete(*self.tree.get_children())
        for row in self.data_rows:
            self.tree.insert("", "end", values=row)
        
        self.search_entry.delete(0, tk.END)
        self._restore_placeholder(None) 
        self._update_status_bar()


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelEditor(root)
    root.mainloop()
